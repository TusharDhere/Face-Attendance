"""
admin/routes/attendance.py
──────────────────────────
Attendance-related routes.

  GET /              → redirect to dashboard
  GET /dashboard     → analytics (summary cards, daily bar chart, low-att table)
  GET /attendance    → paginated attendance log (filter by month)
  GET /export/csv    → download CSV for selected month
  GET /export/excel  → download XLSX with one sheet per month
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Blueprint, render_template, request, Response

from shared.firebase_config import get_firebase

attendance_bp = Blueprint("attendance", __name__)


# ─────────────────────────────────────────────────────────────────────────────
# Firebase helpers
# ─────────────────────────────────────────────────────────────────────────────

def _students() -> dict:
    db, _, _ = get_firebase()
    return db.reference("Students").get() or {}


def _records(month: str | None = None) -> list[dict]:
    """
    Fetch attendance logs from Firestore.
    *month* is 'YYYY-MM'; omit to fetch everything.
    """
    _, _, fs = get_firebase()
    q = fs.collection("attendance")
    if month:
        q = (q.where("date", ">=", f"{month}-01")
              .where("date", "<=", f"{month}-31"))
    out = []
    for doc in q.stream():
        d          = doc.to_dict()
        d["_id"]   = doc.id
        out.append(d)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard  (Phase 2 analytics)
# ─────────────────────────────────────────────────────────────────────────────

@attendance_bp.route("/dashboard")
def dashboard():
    students = _students()
    records  = _records()
    today    = date.today().isoformat()

    total_students = len(students)
    present_today  = len({r["student_id"] for r in records
                          if r.get("date") == today})

    # Per-student totals from RTDB
    all_att = [info.get("total_attendance", 0) for info in students.values()]
    sem_sessions = 30          # assumed sessions in the semester
    if total_students and max(all_att, default=0) > 0:
        avg_pct = int(sum(all_att) / (total_students * sem_sessions) * 100)
    else:
        avg_pct = 0

    # Alerts: students below 75%
    alerts = []
    for sid, info in students.items():
        att = info.get("total_attendance", 0)
        pct = min(int(att / sem_sessions * 100), 100)
        if pct < 75:
            alerts.append({
                "id":    sid,
                "name":  info.get("name",  ""),
                "major": info.get("major", ""),
                "att":   att,
                "pct":   pct,
            })
    alerts.sort(key=lambda a: a["pct"])

    # Daily attendance for last 30 days
    cutoff       = (date.today() - timedelta(days=30)).isoformat()
    daily: dict[str, int] = {}
    for r in records:
        d = r.get("date", "")
        if d and d >= cutoff:
            daily[d] = daily.get(d, 0) + 1

    daily_labels = sorted(daily)[-30:]
    daily_values = [daily[d] for d in daily_labels]

    return render_template(
        "dashboard.html",
        total_students = total_students,
        present_today  = present_today,
        avg_pct        = avg_pct,
        alert_count    = len(alerts),
        alerts         = alerts,
        daily_labels   = daily_labels,
        daily_values   = daily_values,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Attendance log
# ─────────────────────────────────────────────────────────────────────────────

@attendance_bp.route("/attendance")
def attendance_log():
    month   = request.args.get("month", date.today().strftime("%Y-%m"))
    records = _records(month)
    records.sort(
        key=lambda r: (r.get("date", ""), r.get("time", "")),
        reverse=True,
    )
    return render_template("attendance.html", records=records, month=month)


# ─────────────────────────────────────────────────────────────────────────────
# CSV export
# ─────────────────────────────────────────────────────────────────────────────

@attendance_bp.route("/export/csv")
def export_csv():
    month   = request.args.get("month", date.today().strftime("%Y-%m"))
    records = sorted(
        _records(month),
        key=lambda r: (r.get("date", ""), r.get("time", "")),
    )

    label    = datetime.strptime(month, "%Y-%m").strftime("%B_%Y")
    filename = f"Attendance_{label}.csv"

    buf = io.StringIO()
    buf.write("Student ID,Name,Date,Time,Status\n")
    for r in records:
        row = ",".join([
            _esc(r.get("student_id", "")),
            _esc(r.get("name",       "")),
            _esc(r.get("date",       "")),
            _esc(r.get("time",       "")),
            _esc(r.get("status",     "")),
        ])
        buf.write(row + "\n")

    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _esc(v: str) -> str:
    """Minimal CSV escaping: wrap in quotes if value contains comma/quote."""
    v = str(v)
    if "," in v or '"' in v or "\n" in v:
        v = '"' + v.replace('"', '""') + '"'
    return v


@attendance_bp.route("/export/pdf")
def export_pdf():
    """
    Render attendance data as a clean A4-landscape PDF via WeasyPrint.
    Query param: ?month=YYYY-MM  (defaults to current month).
    """
    try:
        from weasyprint import HTML as WP_HTML
    except ImportError:
        return (
            "WeasyPrint is not installed.<br>"
            "Run: <code>pip install weasyprint</code><br>"
            "Linux also needs: "
            "<code>sudo apt install libpango-1.0-0 libcairo2 libpangocairo-1.0-0</code>",
            500,
        )

    month   = request.args.get("month", date.today().strftime("%Y-%m"))
    records = sorted(
        _records(month),
        key=lambda r: (r.get("date", ""), r.get("time", "")),
    )

    label     = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    filename  = f"Attendance_{label.replace(' ', '_')}.pdf"
    generated = datetime.now().strftime("%d %B %Y, %I:%M %p")

    html_str  = render_template(
        "export_pdf.html",
        records   = records,
        month     = label,
        generated = generated,
    )

    pdf_bytes = WP_HTML(string=html_str).write_pdf()

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Excel export  (one sheet per month)
# ─────────────────────────────────────────────────────────────────────────────

@attendance_bp.route("/export/excel")
def export_excel():
    _, _, fs = get_firebase()

    all_records: list[dict] = []
    for doc in fs.collection("attendance").stream():
        all_records.append(doc.to_dict())

    # Group by YYYY-MM
    by_month: dict[str, list[dict]] = {}
    for r in all_records:
        ym = (r.get("date") or "")[:7]
        if ym:
            by_month.setdefault(ym, []).append(r)

    wb  = openpyxl.Workbook()
    wb.remove(wb.active)     # remove default blank sheet

    HDR_FILL  = PatternFill("solid", fgColor="0F172A")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center")
    COLS  = ["Student ID", "Name", "Date", "Time", "Status"]
    COL_W = [14, 24, 12, 10, 10]

    if not by_month:
        # No records yet — add a placeholder so the workbook saves successfully
        ws = wb.create_sheet(title="No Data")
        ws.cell(1, 1, "No attendance records found.")
    else:
        for ym in sorted(by_month):
            label = datetime.strptime(ym, "%Y-%m").strftime("%B %Y")
            ws    = wb.create_sheet(title=label)

            # Header row
            for ci, (col_name, col_w) in enumerate(zip(COLS, COL_W), start=1):
                cell           = ws.cell(row=1, column=ci, value=col_name)
                cell.font      = HDR_FONT
                cell.fill      = HDR_FILL
                cell.alignment = HDR_ALIGN
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(ci)
                ].width = col_w

            ws.row_dimensions[1].height = 22

            # Data rows
            rows = sorted(
                by_month[ym],
                key=lambda r: (r.get("date", ""), r.get("time", "")),
            )
            for ri, r in enumerate(rows, start=2):
                ws.cell(ri, 1, r.get("student_id", ""))
                ws.cell(ri, 2, r.get("name",       ""))
                ws.cell(ri, 3, r.get("date",       ""))
                ws.cell(ri, 4, r.get("time",       ""))
                ws.cell(ri, 5, r.get("status",     ""))

    # Filename uses current month
    filename = f"Attendance_{datetime.now().strftime('%B_%Y')}.xlsx"
    buf      = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
